from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum, Count
from django.http import HttpResponse
from django.utils import timezone
import csv, io

from apps.payments.models import Payment, PaymentStatus
from apps.properties.models import Property, PropertyStatus
from apps.tenants.models import Contract, Tenant
from apps.service_requests.models import ServiceRequest
from apps.users.permissions import IsAdminOrManager


class DashboardStatsView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        total = Property.objects.count()
        return Response({
            "properties": {
                "total":       total,
                "available":   Property.objects.filter(status="available").count(),
                "rented":      Property.objects.filter(status="rented").count(),
            },
            "tenants": {
                "total":            Tenant.objects.filter(status="active").count(),
                "active_contracts": Contract.objects.filter(status="active").count(),
            },
            "payments": {
                "total_collected": float(Payment.objects.aggregate(t=Sum("amount"))["t"] or 0),
                "pending_count":   Payment.objects.filter(status="pending").count(),
                "overdue_count":   Payment.objects.filter(status="overdue").count(),
            },
            "requests": {
                "open": ServiceRequest.objects.filter(
                    status__in=["new","in_progress","waiting"]).count(),
            },
        })


class PaymentAnalyticsView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        year = int(request.query_params.get("year", timezone.now().year))
        monthly = []
        for month in range(1, 13):
            paid = Payment.objects.filter(
                period_year=year, period_month=month, status="paid"
            ).aggregate(t=Sum("amount"))["t"] or 0
            monthly.append({"month": month, "paid": float(paid)})
        return Response({"year": year, "monthly": monthly})


class PropertyAnalyticsView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        return Response({
            "by_status": list(Property.objects.values("status").annotate(count=Count("id"))),
            "by_type":   list(Property.objects.values("property_type__name").annotate(count=Count("id"))),
        })


class RequestAnalyticsView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        return Response({
            "by_status":   list(ServiceRequest.objects.values("status").annotate(count=Count("id"))),
            "by_category": list(ServiceRequest.objects.values("category__name").annotate(count=Count("id"))),
        })


class ExportPaymentsCSVView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        payments = Payment.objects.select_related(
            "contract__tenant__user", "contract__property", "category")
        year = request.query_params.get("year")
        if year:
            payments = payments.filter(period_year=year)

        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="payments.csv"'
        writer = csv.writer(response)
        writer.writerow(["ID","Арендатор","Объект","Категория","Сумма","Статус","Дата","Период"])
        for p in payments:
            writer.writerow([
                p.id,
                p.contract.tenant.user.get_full_name() or p.contract.tenant.user.email,
                p.contract.property.name, p.category.name, p.amount,
                p.get_status_display(), p.due_date.strftime("%d.%m.%Y"),
                f"{p.period_month:02d}/{p.period_year}",
            ])
        return response


class ExportPaymentsExcelView(APIView):
    permission_classes = [IsAdminOrManager]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        payments = Payment.objects.select_related(
            "contract__tenant__user", "contract__property", "category")
        year = request.query_params.get("year")
        if year: payments = payments.filter(period_year=year)

        wb = Workbook()
        ws = wb.active
        ws.title = "Платежи"
        headers = ["ID","Арендатор","Объект","Категория","Сумма","Статус","Дата","Период"]
        hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = hfill
            c.font = Font(color="FFFFFF", bold=True)
            ws.column_dimensions[get_column_letter(i)].width = 20
        for r, p in enumerate(payments, 2):
            ws.cell(r,1,p.id)
            ws.cell(r,2,p.contract.tenant.user.get_full_name())
            ws.cell(r,3,p.contract.property.name)
            ws.cell(r,4,p.category.name)
            ws.cell(r,5,float(p.amount))
            ws.cell(r,6,p.get_status_display())
            ws.cell(r,7,p.due_date.strftime("%d.%m.%Y"))
            ws.cell(r,8,f"{p.period_month:02d}/{p.period_year}")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
        return response

